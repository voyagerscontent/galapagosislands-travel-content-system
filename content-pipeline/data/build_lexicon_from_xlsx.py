#!/usr/bin/env python3
"""Build data/lexicon.json from the High-Entropy Injectors (BCP) workbook.

The workbook's "All Words" sheet has columns: Word, Word Type, Category, Group.
Group is exactly destination / activity / feeling. We compile per-(group,
category) POS pools that lexical_injector uses to swap generic adjectives/verbs
for contextually-mapped high-BCP words. Re-run when the workbook changes:

  python content-pipeline/data/build_lexicon_from_xlsx.py "/path/to/High-Entropy Injectors (2.5-4.0 BPC).xlsx"
"""
import json
import os
import sys
import collections

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "lexicon.json")

# Generic, high-probability words the injector targets for replacement.
GENERIC_ADJECTIVES = [
    "beautiful", "stunning", "amazing", "incredible", "breathtaking", "gorgeous",
    "lovely", "nice", "great", "wonderful", "spectacular", "magnificent", "majestic",
    "pristine", "lush", "vast", "remote", "peaceful", "serene", "tranquil", "dramatic",
    "rugged", "vibrant", "memorable", "unforgettable", "picturesque", "scenic",
    "impressive", "striking", "charming", "idyllic", "pretty", "beautifull",
]
GENERIC_VERBS = [
    "see", "explore", "enjoy", "experience", "discover", "offer", "witness", "watch",
    "visit", "find", "observe", "encounter", "spot", "provide", "feel", "take in",
]

# WordType -> pool bucket
TYPE_MAP = {"Adjective": "adjective", "Sensory": "sensory", "Verb": "verb",
            "Noun": "noun", "Place": "place"}
GROUP_MAP = {"Destination": "destination", "Activity": "activity", "Feeling": "feeling"}


import re as _re

# Salt sources appropriate for GALAPAGOS content (generic travel/cruise + Galapagos).
# Other-destination + off-topic sources are excluded so we never splice, e.g., a UK
# car-park line or a South Georgia penguin line into a Galapagos article.
# GALAPAGOS-SPECIFIC sources only. General "Cruise"/"travel" forums carry
# Mediterranean/Alaska/polar cruise talk, so they are excluded; the allow-list is
# Galapagos threads + Galapagos operators/blogs (Quasar, Pikaia, Lindblad, etc.).
_SALT_ALLOW = _re.compile(r"galapagos|quasar|pikaia|lindblad|nathab|thinkgalapagos|"
                          r"metropolitan|happygringo|itchyfeet|andysworld|laidback|"
                          r"ciaobambino|bucketlist|roadto197|suewhere|velvetescape|"
                          r"globetrotter|shoestring", _re.I)
# Content-level safety net: drop any sentence naming a non-Galapagos place/animal,
# incl. polar wildlife/terms that leak in via general/expedition cruise forums.
_SALT_DENY = _re.compile(
    r"\b(vietnam|danang|hoi an|patagonia|antarctic|antarctica|south georgia|falkland|"
    r"cusco|machu|iquitos|loreto|brasilia|sao paolo|sao paulo|el calafate|chile|"
    r"argentina|peru|brazil|south africa|car park|parking|pod |terminal|shuttle|"
    r"skua|iceberg|glacier|tundra|gentoo|adelie|chinstrap|king penguin|emperor|krill|"
    r"orca|humpback|leopard seal|elephant seal|peninsula|moss|types of penguin|"
    r"\d+\s+(?:types|different)\s+.*penguin|hundreds of penguin|100s of penguin)", _re.I)


def load_salt(xlsx: str, cap: int = 1200) -> list:
    """Load verbatim human 'salt' sentences from the master travel-phrases workbook,
    filtered to complete, Galapagos-appropriate lines and tagged by source."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["phrases"]
    seen, out = set(), []
    for row in ws.iter_rows(min_row=2, values_only=True):
        txt = (str(row[0]).strip() if row[0] else "")
        src = str(row[4] or "")
        if not txt or txt.lower() in seen:
            continue
        wc = len(txt.split())
        if not (4 <= wc <= 14):                       # complete-but-short
            continue
        if not txt[0].isupper() or not txt.rstrip().endswith((".", "!", "?")):
            continue                                   # must be a complete sentence
        if "http" in txt or "@" in txt or "/" in txt:
            continue
        if not _SALT_ALLOW.search(src) or _SALT_DENY.search(txt):
            continue
        seen.add(txt.lower())
        low = src.lower()
        tags = ["general"]
        if "galapagos" in low:
            tags += ["galapagos", "wildlife", "destination"]
        if _re.search(r"cruise|ship|expedition|quasar|pikaia|lindblad", low):
            tags += ["cruise", "activity"]
        out.append({"text": txt, "kind": "phrase", "tags": sorted(set(tags))})
        if len(out) >= cap:
            break
    return out


def load_long(xlsx: str, cap: int = 900) -> list:
    """Load LONGER verbatim salt sentences (col 'sentence', col 'parameter'=topic),
    filtered to Galapagos-relevant complete sentences (8-30 words)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    seen, out = set(), []
    for row in ws.iter_rows(min_row=2, values_only=True):
        txt = (str(row[0]).strip() if row[0] else "")
        param = str(row[2] or "").lower()
        if not txt or txt.lower() in seen:
            continue
        if "galapagos" not in param and "expedition cruise" not in param:
            continue                                   # Galapagos-relevant topics only
        wc = len(txt.split())
        if not (8 <= wc <= 30):
            continue
        if not txt[0].isupper() or not txt.rstrip().endswith((".", "!", "?")):
            continue
        if "http" in txt or _SALT_DENY.search(txt):
            continue
        seen.add(txt.lower())
        out.append({"text": txt, "kind": "long",
                    "tags": ["galapagos", "wildlife", "destination", "cruise", "general"]})
        if len(out) >= cap:
            break
    return out


def main(xlsx: str, salt_xlsx: str = None, long_xlsx: str = None):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["All Words"]
    pools = {"destination": {}, "activity": {}, "feeling": {}}
    counts = collections.Counter()
    for word, wtype, category, group in ws.iter_rows(min_row=2, values_only=True):
        if not (word and wtype and category and group):
            continue
        g = GROUP_MAP.get(str(group).strip())
        bucket = TYPE_MAP.get(str(wtype).strip())
        if not g or not bucket:
            continue
        cat = str(category).strip()
        w = str(word).strip()
        pools[g].setdefault(cat, {"adjective": [], "sensory": [], "verb": [], "noun": [], "place": []})
        if w not in pools[g][cat][bucket]:
            pools[g][cat][bucket].append(w)
            counts[bucket] += 1

    out = {
        "_meta": {
            "name": "high_bcp_lexicon",
            "description": "High-BCP (2.5-4.0 BPC) travel vocabulary compiled from the "
                           "High-Entropy Injectors workbook, categorized by destination / "
                           "activity / feeling and POS. Used by lexical_injector to swap generic "
                           "adjectives/verbs for contextually-mapped low-probability words.",
            "source": os.path.basename(xlsx),
            "version": "1.0.0",
            "totals": dict(counts),
        },
        "generic_adjectives": GENERIC_ADJECTIVES,
        "generic_verbs": GENERIC_VERBS,
        "pools": pools,
        # verbatim human "salt" sentences from the master travel-phrases workbook
        # (Galapagos-appropriate, complete, tagged). Falls back to a small seed.
        "human_sentences": ((load_salt(salt_xlsx) if salt_xlsx else []) +
                            (load_long(long_xlsx) if long_xlsx else [])) or [
            {"text": "You forget your phone exists out there, and it takes a day to notice.", "kind": "phrase", "tags": ["general"]},
        ],
    }
    import collections as _c
    out["_meta"]["salt_counts"] = dict(_c.Counter(s.get("kind","phrase") for s in out["human_sentences"]))
    with open(OUT, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)
    ncats = sum(len(v) for v in pools.values())
    print(f"wrote {OUT}: {ncats} categories, pools {dict(counts)}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("usage: build_lexicon_from_xlsx.py <xlsx path>")
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None, sys.argv[3] if len(sys.argv) > 3 else None)
